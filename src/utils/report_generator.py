#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Module de génération de rapports pour RedForge
Génère des rapports en différents formats avec support furtif et APT
Version avec templates, export avancé et métriques
"""

import json
import csv
import os
import zipfile
from datetime import datetime
from typing import Dict, Any, List, Optional, Union
from pathlib import Path
from dataclasses import dataclass, field
from enum import Enum

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from src.utils.color_output import console


class ReportFormat(Enum):
    """Formats de rapport disponibles"""
    JSON = "json"
    CSV = "csv"
    HTML = "html"
    PDF = "pdf"
    TXT = "txt"
    MARKDOWN = "md"
    EXCEL = "xlsx"
    ZIP = "zip"


@dataclass
class ReportConfig:
    """Configuration du rapport"""
    title: str = "RedForge - Rapport de Pentest"
    include_timestamp: bool = True
    include_summary: bool = True
    include_details: bool = True
    include_recommendations: bool = True
    include_evidence: bool = True
    max_vulnerabilities: int = 0  # 0 = illimité
    sort_by: str = "severity"  # severity, type, date
    template: str = "default"
    stealth_mode: bool = False
    apt_mode: bool = False


class ReportGenerator:
    """Générateur de rapports avancé avec support furtif"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.config = ReportConfig()
        self._setup_custom_styles()
    
    def set_config(self, config: ReportConfig):
        """Configure le générateur de rapports"""
        self.config = config
    
    def set_stealth_mode(self, enabled: bool):
        """Active/désactive le mode furtif"""
        self.config.stealth_mode = enabled
    
    def set_apt_mode(self, enabled: bool):
        """Active/désactive le mode APT"""
        self.config.apt_mode = enabled
    
    def _setup_custom_styles(self):
        """Configure les styles personnalisés"""
        # Titre principal
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#d32f2f'),
            spaceAfter=30,
            alignment=1  # Centré
        ))
        
        # Sous-titre
        self.styles.add(ParagraphStyle(
            name='CustomSubtitle',
            parent=self.styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#666666'),
            spaceAfter=20,
            alignment=1
        ))
        
        # En-tête de section
        self.styles.add(ParagraphStyle(
            name='CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#333333'),
            spaceAfter=12,
            spaceBefore=20
        ))
        
        # Titre de vulnérabilité
        self.styles.add(ParagraphStyle(
            name='VulnerabilityTitle',
            parent=self.styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#d32f2f'),
            fontName='Helvetica-Bold',
            spaceAfter=6
        ))
        
        # Métadonnées
        self.styles.add(ParagraphStyle(
            name='Metadata',
            parent=self.styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#888888'),
            alignment=1
        ))
    
    def generate(self, data: Dict[str, Any], output_file: str, 
                 format: Union[str, ReportFormat] = ReportFormat.HTML) -> bool:
        """
        Génère un rapport dans le format spécifié
        
        Args:
            data: Données du rapport
            output_file: Fichier de sortie
            format: Format du rapport
        """
        if isinstance(format, str):
            format = ReportFormat(format.lower())
        
        methods = {
            ReportFormat.JSON: self.generate_json,
            ReportFormat.CSV: self.generate_csv,
            ReportFormat.HTML: self.generate_html,
            ReportFormat.PDF: self.generate_pdf,
            ReportFormat.TXT: self.generate_txt,
            ReportFormat.MARKDOWN: self.generate_markdown,
            ReportFormat.EXCEL: self.generate_excel,
            ReportFormat.ZIP: self.generate_zip
        }
        
        if format not in methods:
            console.print_error(f"Format non supporté: {format.value}")
            return False
        
        # Mode APT: réduire les détails
        if self.config.apt_mode:
            self.config.include_evidence = False
            self.config.max_vulnerabilities = 50
        
        return methods[format](data, output_file)
    
    def generate_json(self, data: Dict[str, Any], output_file: str) -> bool:
        """Génère un rapport JSON"""
        try:
            report = {
                'generated_at': datetime.now().isoformat(),
                'tool': 'RedForge',
                'version': '2.0.0',
                'config': {
                    'stealth_mode': self.config.stealth_mode,
                    'apt_mode': self.config.apt_mode
                },
                'summary': self._generate_summary(data),
                'data': data
            }
            
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(report, f, indent=2, ensure_ascii=False, default=str)
            
            if not self.config.stealth_mode:
                console.print_success(f"Rapport JSON généré: {output_file}")
            return True
            
        except Exception as e:
            console.print_error(f"Erreur génération JSON: {e}")
            return False
    
    def generate_csv(self, data: Dict[str, Any], output_file: str) -> bool:
        """Génère un rapport CSV"""
        try:
            vulnerabilities = self._extract_vulnerabilities(data)
            
            if not vulnerabilities:
                console.print_warning("Aucune vulnérabilité à exporter")
                return False
            
            # Limiter le nombre
            if self.config.max_vulnerabilities > 0:
                vulnerabilities = vulnerabilities[:self.config.max_vulnerabilities]
            
            with open(output_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=vulnerabilities[0].keys())
                writer.writeheader()
                writer.writerows(vulnerabilities)
            
            if not self.config.stealth_mode:
                console.print_success(f"Rapport CSV généré: {output_file}")
            return True
            
        except Exception as e:
            console.print_error(f"Erreur génération CSV: {e}")
            return False
    
    def generate_html(self, data: Dict[str, Any], output_file: str) -> bool:
        """Génère un rapport HTML"""
        try:
            html = self._build_html_report(data)
            
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(html)
            
            if not self.config.stealth_mode:
                console.print_success(f"Rapport HTML généré: {output_file}")
            return True
            
        except Exception as e:
            console.print_error(f"Erreur génération HTML: {e}")
            return False
    
    def generate_pdf(self, data: Dict[str, Any], output_file: str) -> bool:
        """Génère un rapport PDF"""
        try:
            doc = SimpleDocTemplate(output_file, pagesize=A4)
            story = []
            
            # Titre
            title = Paragraph(self.config.title, self.styles['CustomTitle'])
            story.append(title)
            story.append(Spacer(1, 0.2 * inch))
            
            # Date
            if self.config.include_timestamp:
                date_str = f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
                story.append(Paragraph(date_str, self.styles['Metadata']))
                story.append(Spacer(1, 0.3 * inch))
            
            # Résumé
            if self.config.include_summary:
                story.append(Paragraph("Résumé des vulnérabilités", self.styles['CustomHeading']))
                summary = self._generate_summary(data)
                
                # Tableau de résumé
                summary_data = [[k, str(v)] for k, v in summary.items()]
                summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(summary_table)
                story.append(Spacer(1, 0.3 * inch))
            
            # Vulnérabilités
            if self.config.include_details:
                story.append(Paragraph("Détail des vulnérabilités", self.styles['CustomHeading']))
                vulnerabilities = self._extract_vulnerabilities(data)
                
                # Limiter le nombre
                if self.config.max_vulnerabilities > 0:
                    vulnerabilities = vulnerabilities[:self.config.max_vulnerabilities]
                
                for vuln in vulnerabilities:
                    story.append(Paragraph(f"<b>{vuln.get('type', 'Unknown')}</b>", 
                                          self.styles['VulnerabilityTitle']))
                    story.append(Paragraph(f"Sévérité: {vuln.get('severity', 'N/A')}", 
                                          self.styles['Normal']))
                    story.append(Paragraph(f"Paramètre: {vuln.get('parameter', 'N/A')}", 
                                          self.styles['Normal']))
                    story.append(Paragraph(f"Détails: {vuln.get('details', 'N/A')}", 
                                          self.styles['Normal']))
                    
                    if self.config.include_evidence and vuln.get('evidence'):
                        story.append(Paragraph(f"Preuve: {vuln['evidence'][:200]}", 
                                              self.styles['Normal']))
                    
                    story.append(Spacer(1, 0.1 * inch))
            
            doc.build(story)
            
            if not self.config.stealth_mode:
                console.print_success(f"Rapport PDF généré: {output_file}")
            return True
            
        except Exception as e:
            console.print_error(f"Erreur génération PDF: {e}")
            return False
    
    def generate_txt(self, data: Dict[str, Any], output_file: str) -> bool:
        """Génère un rapport texte"""
        try:
            lines = []
            lines.append("=" * 70)
            lines.append(self.config.title)
            lines.append("=" * 70)
            lines.append(f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
            lines.append("")
            
            # Résumé
            if self.config.include_summary:
                lines.append("RÉSUMÉ DES VULNÉRABILITÉS")
                lines.append("-" * 40)
                summary = self._generate_summary(data)
                for key, value in summary.items():
                    lines.append(f"{key}: {value}")
                lines.append("")
            
            # Vulnérabilités
            if self.config.include_details:
                lines.append("DÉTAIL DES VULNÉRABILITÉS")
                lines.append("-" * 40)
                vulnerabilities = self._extract_vulnerabilities(data)
                
                if self.config.max_vulnerabilities > 0:
                    vulnerabilities = vulnerabilities[:self.config.max_vulnerabilities]
                
                for i, vuln in enumerate(vulnerabilities, 1):
                    lines.append(f"\n{i}. {vuln.get('type', 'Unknown')}")
                    lines.append(f"   Sévérité: {vuln.get('severity', 'N/A')}")
                    lines.append(f"   Paramètre: {vuln.get('parameter', 'N/A')}")
                    lines.append(f"   Détails: {vuln.get('details', 'N/A')}")
            
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write('\n'.join(lines))
            
            if not self.config.stealth_mode:
                console.print_success(f"Rapport texte généré: {output_file}")
            return True
            
        except Exception as e:
            console.print_error(f"Erreur génération texte: {e}")
            return False
    
    def generate_markdown(self, data: Dict[str, Any], output_file: str) -> bool:
        """Génère un rapport Markdown"""
        try:
            lines = []
            lines.append(f"# {self.config.title}")
            lines.append("")
            lines.append(f"*Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}*")
            lines.append("")
            
            # Résumé
            if self.config.include_summary:
                lines.append("## Résumé des vulnérabilités")
                lines.append("")
                summary = self._generate_summary(data)
                for key, value in summary.items():
                    lines.append(f"- **{key}**: {value}")
                lines.append("")
            
            # Vulnérabilités
            if self.config.include_details:
                lines.append("## Détail des vulnérabilités")
                lines.append("")
                vulnerabilities = self._extract_vulnerabilities(data)
                
                if self.config.max_vulnerabilities > 0:
                    vulnerabilities = vulnerabilities[:self.config.max_vulnerabilities]
                
                for vuln in vulnerabilities:
                    severity = vuln.get('severity', 'unknown').lower()
                    severity_icon = "🔴" if severity == "critical" else "🟠" if severity == "high" else "🟡" if severity == "medium" else "🔵"
                    lines.append(f"### {severity_icon} {vuln.get('type', 'Unknown')}")
                    lines.append(f"- **Sévérité**: {vuln.get('severity', 'N/A')}")
                    lines.append(f"- **Paramètre**: {vuln.get('parameter', 'N/A')}")
                    lines.append(f"- **Détails**: {vuln.get('details', 'N/A')}")
                    lines.append("")
            
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write('\n'.join(lines))
            
            if not self.config.stealth_mode:
                console.print_success(f"Rapport Markdown généré: {output_file}")
            return True
            
        except Exception as e:
            console.print_error(f"Erreur génération Markdown: {e}")
            return False
    
    def generate_excel(self, data: Dict[str, Any], output_file: str) -> bool:
        """Génère un rapport Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            
            # Feuille de résumé
            ws_summary = wb.active
            ws_summary.title = "Résumé"
            
            # En-têtes
            ws_summary['A1'] = self.config.title
            ws_summary['A1'].font = Font(size=16, bold=True)
            ws_summary['A3'] = f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            
            if self.config.include_summary:
                ws_summary['A5'] = "Résumé des vulnérabilités"
                ws_summary['A5'].font = Font(bold=True, size=12)
                
                row = 7
                summary = self._generate_summary(data)
                for key, value in summary.items():
                    ws_summary[f'A{row}'] = key
                    ws_summary[f'B{row}'] = value
                    row += 1
            
            # Feuille des vulnérabilités
            if self.config.include_details:
                ws_vulns = wb.create_sheet("Vulnérabilités")
                
                # En-têtes
                headers = ['Type', 'Sévérité', 'Paramètre', 'Détails', 'Preuve']
                for col, header in enumerate(headers, 1):
                    cell = ws_vulns.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                vulnerabilities = self._extract_vulnerabilities(data)
                
                if self.config.max_vulnerabilities > 0:
                    vulnerabilities = vulnerabilities[:self.config.max_vulnerabilities]
                
                for row, vuln in enumerate(vulnerabilities, 2):
                    ws_vulns.cell(row=row, column=1, value=vuln.get('type', ''))
                    ws_vulns.cell(row=row, column=2, value=vuln.get('severity', ''))
                    ws_vulns.cell(row=row, column=3, value=vuln.get('parameter', ''))
                    ws_vulns.cell(row=row, column=4, value=vuln.get('details', ''))
                    ws_vulns.cell(row=row, column=5, value=vuln.get('evidence', '')[:500])
            
            wb.save(output_file)
            
            if not self.config.stealth_mode:
                console.print_success(f"Rapport Excel généré: {output_file}")
            return True
            
        except ImportError:
            console.print_error("openpyxl non installé. Installez-le avec: pip install openpyxl")
            return False
        except Exception as e:
            console.print_error(f"Erreur génération Excel: {e}")
            return False
    
    def generate_zip(self, data: Dict[str, Any], output_file: str) -> bool:
        """Génère un rapport ZIP contenant plusieurs formats"""
        try:
            temp_dir = Path(output_file).parent / f"temp_report_{int(datetime.now().timestamp())}"
            temp_dir.mkdir(exist_ok=True)
            
            formats = [ReportFormat.JSON, ReportFormat.HTML, ReportFormat.CSV, ReportFormat.TXT]
            generated_files = []
            
            for fmt in formats:
                fmt_file = temp_dir / f"report.{fmt.value}"
                if self.generate(data, str(fmt_file), fmt):
                    generated_files.append(fmt_file)
            
            with zipfile.ZipFile(output_file, 'w', zipfile.ZIP_DEFLATED) as zf:
                for file in generated_files:
                    zf.write(file, file.name)
            
            # Nettoyage
            import shutil
            shutil.rmtree(temp_dir)
            
            if not self.config.stealth_mode:
                console.print_success(f"Rapport ZIP généré: {output_file}")
            return True
            
        except Exception as e:
            console.print_error(f"Erreur génération ZIP: {e}")
            return False
    
    def _build_html_report(self, data: Dict[str, Any]) -> str:
        """Construit le rapport HTML"""
        summary = self._generate_summary(data)
        vulnerabilities = self._extract_vulnerabilities(data)
        
        # Limiter le nombre
        if self.config.max_vulnerabilities > 0:
            vulnerabilities = vulnerabilities[:self.config.max_vulnerabilities]
        
        # Construction des lignes de vulnérabilités
        vuln_rows = ""
        for vuln in vulnerabilities:
            severity = vuln.get('severity', 'medium').lower()
            severity_class = "critical" if severity == "critical" else "high" if severity == "high" else "medium" if severity == "medium" else "low"
            vuln_rows += f"""
            <tr class="severity-{severity_class}">
                <td>{vuln.get('type', 'Unknown')}</td>
                <td><span class="badge badge-{severity_class}">{vuln.get('severity', 'N/A')}</span></td>
                <td>{vuln.get('parameter', 'N/A')}</td>
                <td>{vuln.get('details', 'N/A')[:100]}</td>
            </tr>
            """
        
        html = f"""<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    <title>{self.config.title}</title>
    <style>
        body {{
            font-family: 'Segoe UI', Arial, sans-serif;
            margin: 20px;
            background: #f5f5f5;
        }}
        .container {{
            max-width: 1200px;
            margin: auto;
            background: white;
            padding: 30px;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        h1 {{ color: #d32f2f; border-left: 4px solid #d32f2f; padding-left: 15px; }}
        h2 {{ color: #333; border-bottom: 2px solid #eee; padding-bottom: 10px; }}
        .summary {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
            gap: 15px;
            margin: 20px 0;
        }}
        .summary-card {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 10px;
            text-align: center;
            transition: transform 0.3s;
        }}
        .summary-card:hover {{
            transform: translateY(-5px);
        }}
        .summary-value {{
            font-size: 32px;
            font-weight: bold;
        }}
        .summary-label {{
            font-size: 14px;
            opacity: 0.9;
            margin-top: 5px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }}
        th, td {{
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #ddd;
        }}
        th {{
            background: #f8f9fa;
            font-weight: bold;
            color: #333;
        }}
        tr:hover {{
            background: #f5f5f5;
        }}
        .severity-critical {{ background-color: #ffebee; }}
        .severity-high {{ background-color: #fff3e0; }}
        .severity-medium {{ background-color: #e3f2fd; }}
        .severity-low {{ background-color: #e8f5e9; }}
        .badge {{
            display: inline-block;
            padding: 3px 10px;
            border-radius: 20px;
            font-size: 11px;
            font-weight: bold;
        }}
        .badge-critical {{ background: #7b1fa2; color: white; }}
        .badge-high {{ background: #d32f2f; color: white; }}
        .badge-medium {{ background: #ff9800; color: white; }}
        .badge-low {{ background: #4caf50; color: white; }}
        .footer {{
            margin-top: 30px;
            padding-top: 20px;
            border-top: 1px solid #ddd;
            text-align: center;
            font-size: 12px;
            color: #666;
        }}
        .mode-badge {{
            display: inline-block;
            padding: 2px 8px;
            border-radius: 4px;
            font-size: 10px;
            margin-left: 10px;
        }}
        .mode-stealth {{ background: #607d8b; color: white; }}
        .mode-apt {{ background: #9c27b0; color: white; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🔴 {self.config.title}
            {"<span class='mode-badge mode-stealth'>Stealth</span>" if self.config.stealth_mode else ""}
            {"<span class='mode-badge mode-apt'>APT</span>" if self.config.apt_mode else ""}
        </h1>
        <p>Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</p>
        
        <div class="summary">
            <div class="summary-card" style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);">
                <div class="summary-value">{summary.get('Total vulnérabilités', 0)}</div>
                <div class="summary-label">Vulnérabilités</div>
            </div>
            <div class="summary-card" style="background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);">
                <div class="summary-value">{summary.get('Critique', 0)}</div>
                <div class="summary-label">Critique</div>
            </div>
            <div class="summary-card" style="background: linear-gradient(135deg, #fa709a 0%, #fee140 100%);">
                <div class="summary-value">{summary.get('Élevée', 0)}</div>
                <div class="summary-label">Élevée</div>
            </div>
            <div class="summary-card" style="background: linear-gradient(135deg, #a8edea 0%, #fed6e3 100%);">
                <div class="summary-value">{summary.get('Moyenne', 0)}</div>
                <div class="summary-label">Moyenne</div>
            </div>
            <div class="summary-card" style="background: linear-gradient(135deg, #d4fc79 0%, #96e6a1 100%);">
                <div class="summary-value">{summary.get('Basse', 0)}</div>
                <div class="summary-label">Basse</div>
            </div>
        </div>
        
        <h2>📋 Liste des vulnérabilités</h2>
        <table>
            <thead>
                <tr>
                    <th>Type</th>
                    <th>Sévérité</th>
                    <th>Paramètre</th>
                    <th>Détails</th>
                </tr>
            </thead>
            <tbody>
                {vuln_rows if vuln_rows else '<tr><td colspan="4">Aucune vulnérabilité détectée</td></tr>'}
            </tbody>
        </table>
        
        <div class="footer">
            <p>Rapport généré par RedForge v2.0.0 - Plateforme d'Orchestration de Pentest Web</p>
            <p>Mode furtif: {'Activé' if self.config.stealth_mode else 'Désactivé'} | Mode APT: {'Activé' if self.config.apt_mode else 'Désactivé'}</p>
        </div>
    </div>
</body>
</html>"""
        
        return html
    
    def _generate_summary(self, data: Dict[str, Any]) -> Dict[str, int]:
        """Génère un résumé des vulnérabilités"""
        summary = {
            'Total vulnérabilités': 0,
            'Critique': 0,
            'Élevée': 0,
            'Moyenne': 0,
            'Basse': 0
        }
        
        vulnerabilities = self._extract_vulnerabilities(data)
        summary['Total vulnérabilités'] = len(vulnerabilities)
        
        for vuln in vulnerabilities:
            severity = vuln.get('severity', '').upper()
            if severity == 'CRITICAL':
                summary['Critique'] += 1
            elif severity == 'HIGH':
                summary['Élevée'] += 1
            elif severity == 'MEDIUM':
                summary['Moyenne'] += 1
            elif severity == 'LOW':
                summary['Basse'] += 1
        
        return summary
    
    def _extract_vulnerabilities(self, data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Extrait toutes les vulnérabilités des données"""
        vulnerabilities = []
        
        def extract(obj):
            if isinstance(obj, dict):
                if 'vulnerabilities' in obj and isinstance(obj['vulnerabilities'], list):
                    vulnerabilities.extend(obj['vulnerabilities'])
                for value in obj.values():
                    extract(value)
            elif isinstance(obj, list):
                for item in obj:
                    extract(item)
        
        extract(data)
        
        # Trier par sévérité
        severity_order = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3}
        vulnerabilities.sort(key=lambda x: severity_order.get(x.get('severity', 'LOW').upper(), 4))
        
        return vulnerabilities


# Instance globale
report_gen = ReportGenerator()


if __name__ == "__main__":
    test_data = {
        'sql_injection': {
            'vulnerabilities': [
                {'type': 'SQL Injection', 'severity': 'CRITICAL', 'parameter': 'id', 'details': 'Injection SQL dans le paramètre id', 'evidence': 'Payload: 1\' OR \'1\'=\'1'},
                {'type': 'XSS', 'severity': 'HIGH', 'parameter': 'q', 'details': 'Reflected XSS', 'evidence': 'Payload: <script>alert(1)</script>'}
            ]
        }
    }
    
    # Test des différents formats
    report_gen.generate_html(test_data, 'test_report.html')
    report_gen.generate_json(test_data, 'test_report.json')
    report_gen.generate_csv(test_data, 'test_report.csv')
    report_gen.generate_txt(test_data, 'test_report.txt')
    report_gen.generate_markdown(test_data, 'test_report.md')
    
    # Mode APT
    report_gen.set_apt_mode(True)
    report_gen.generate_html(test_data, 'test_report_apt.html')
    
    console.print_success("Tests terminés")